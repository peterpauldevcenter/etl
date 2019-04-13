"""
The historical data is saved in a likert-like format on the Student Roster file.

Each year is captured under a different column, and so the ranges must be mapped for ETL.
"""
import re
from functools import partial, reduce
from typing import Iterable, List, Tuple, Dict, AnyStr, Any

from etl.excel_ingestion.youthsurvey.new import YouthSurveyValidationException
from etl.excel_ingestion.youthsurvey.contrib import add_to_set_indicate_size_change
from collections import namedtuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

question_re = re.compile(r'(?P<firstword>\A\w*)(?=\s)?.*(?P<year>\d{4})\s?(?P<question>.*)')
html_re = re.compile(r'<.*?>')

QuestionNamedTuple = namedtuple('QuestionNamedTuple', 'index column_label season year question')
"""A record to hold metadata about a question.

Args:
    index: A column's index
    column_label: The excel label (for debugging) like AA or AB, etc.
    season: The season this column's data belongs to
    year: The year this column belongs to
    question: The question this column belongs to
"""
ResultNamedTuple = namedtuple('ResultNamedTuple', 'season year question answer')
"""A record to hold an answer for a each question per season per year."""


def get_workbook(path):
    return load_workbook(path, read_only=True, data_only=True)


def get_data(workbook):
    return list(workbook['Students'].values)


def validate_and_get_question_metadata(data):
    """Validates the Excel column structure and returns question metadata records."""
    headers = data[0]
    if headers[0].lower() not in ('markelid', 'studentid', 'id', 'pk'):
        raise YouthSurveyValidationException(
            'First column is not a known student id alias. Make sure to add student '
            'id as the first column with a header like MarkelID, StudentID, ID, or PK.'
        )

    question_data = get_question_named_tuple_from_headers(headers)
    clean, duplicates = check_duplicates(question_data)

    if duplicates:
        raise YouthSurveyValidationException(
            f'Duplicate question metadata found: \n {duplicates}'
        )

    all_fall_questions = get_season_questions('fall', clean)
    all_spring_questions = get_season_questions('spring', clean)

    check_headers(all_fall_questions)
    check_headers(all_spring_questions)
    return clean


def get_student_results(question_metadata: Iterable[QuestionNamedTuple],
                        row) -> 'Dict[AnyStr, Dict[AnyStr: Dict[AnyStr, Any]]]':
    """Processes a row in the data set returning the students data.

    Args:
        question_metadata: The collection of QuestionNamedTuples
        row: The row to process

    Returns:
        A dictionary whose key is the student id, and value is another dictionary where each
        key '{season} {year}' contains a list of the questions answers for that biannual.

        It will look like ``{student_id: {'spring 2018': {question: answer}`` etc

        Season will be lowercase.
    """
    student_token = str(row[0])

    all_semesters = {}
    ret = {student_token: all_semesters}

    for meta in question_metadata:
        key = f'{meta.season.lower()} {meta.year}'
        all_semesters.setdefault(key, {})
        answer = row[meta.index]
        all_semesters[key].update({meta.question: answer})
    return ret


def get_season_year_question_from_header(header_value: str) -> dict:
    """Extracts the values of "firstword", "year", and "question", from the header as a dict.

    Args:
        header_value:

    Raises:
        :exc:`YouthSurveyValidationException`: When the regex doesn't match or not all the keys are populated.
    """
    match = question_re.match(header_value)
    if not match:
        raise YouthSurveyValidationException(
            f'Could not match regular expression format of header {header_value}.'
        )
    data = match.groupdict()
    for key, value in data.items():
        if not value:
            raise YouthSurveyValidationException(
                f'Missing {key} for header {header_value}.'
            )
    return data


def get_question_named_tuple_from_headers(headers: Iterable[str]) -> List[QuestionNamedTuple]:
    """Applies :func:`get_season_year_question_from_header` to each element in headers.

    Creates a :class:`QuestionNamedTuple` metadata object and adds it to the return collection when found.

    Args:
        headers: An iterable of data headers
    """
    named_tuples = []
    for index, value in enumerate(headers):
        try:
            data = get_season_year_question_from_header(value)
            firstword, year, question = map(data.get, 'firstword year question'.split())
            # I couldn't get the regex to be precise enough to filter on descriptionreading and description math so...
            if firstword.lower() in ('spring', 'fall') and question.lower() not in (
                    'descriptionreading', 'descriptionmath'):
                # Strip random html tags from the question
                question = html_re.sub('', question)
                column_label = get_column_letter(index + 1)
                named_tuples.append(QuestionNamedTuple(index, column_label, firstword, year, question))
        except YouthSurveyValidationException:
            continue
    return named_tuples


def normalize_fields(tuple_):
    return tuple((item.lower() for item in tuple_))


def check_duplicates(nt_data) -> Tuple[List, List]:
    """Metadata records must be unique by season, year, question.

    Returns a tuple of two collections: clean, dirty
    """
    check = set()
    ret = []
    duplicates = []

    for record in nt_data:
        normalized = normalize_fields(record[2:])
        _, added = add_to_set_indicate_size_change(check, normalized)
        if added:
            ret.append(record)
        else:
            duplicates.append(record)
    return ret, duplicates


def get_questions_for_year(question_data, year):
    year = str(year)
    return [record for record in question_data if record.year == year]


def get_data_for_year_season(question_data, year, season):
    filtered = get_questions_for_year(question_data, year)
    return [record for record in filtered if record.season.lower() == season.lower()]


def get_season_questions(season: str,
                         question_data: Iterable[QuestionNamedTuple]):
    """Group season questions by year for all the questions for years 2015-2018.

    Args:
        season: either "spring" or "fall" case insensitive
        question_data: The list of records
    """
    data_getter = partial(get_data_for_year_season, question_data, season=season)
    # noinspection PyTypeChecker
    return tuple(map(data_getter, [2015, 2016, 2017, 2018]))


def check_headers(year_season_collection: Iterable[QuestionNamedTuple]):
    """Checks to make sure all the questions are present, and their text content in the correct format and
    is identical for each biannual.

    Passing the test means we can safely rely on the question header position to get the data.

    Args:
        year_season_collection: An iterable of :obj:`QuestionNamedTuple` whose seasons are the same.
    """

    def my_reducer(year_season_data, next_year_season_data):
        year_record, next_year_record = year_season_data[0], next_year_season_data[0]
        first_collection_year, second_collection_year = year_record.year, next_year_record.year
        first_collection_season, second_collection_season = year_record.season, next_year_record.season

        err_msg_stem = (f'{first_collection_season} {first_collection_year} and '
                        f'{second_collection_season} {second_collection_year} question collections')

        accu_len, next_item_len = len(year_season_data), len(next_year_season_data)
        if accu_len != next_item_len:
            raise YouthSurveyValidationException(
                f'{err_msg_stem} are of different length!'
            )

        season_q_text = set((record.question for record in year_season_data))
        next_season_q_text = set((record.question for record in next_year_season_data))

        different_text = season_q_text - next_season_q_text
        if different_text:
            raise YouthSurveyValidationException(
                f'{err_msg_stem} contain different question text values. \n'
                '\n'.join([q_text for q_text in different_text])
            )
        return next_year_season_data

    reduce(my_reducer, year_season_collection)
